import openpyxl as xls
from openpyxl.chart import BarChart, Reference

wb = xls.load_workbook('transactions.xlsx')
trasnactions = wb['transaction1']


# Working with rows of an excel sheet
for row in range(1, trasnactions.max_row + 1):
    print(row)

# Getting all rows of cell 3 
for row in range(1, trasnactions.max_row + 1):
    cell = trasnactions.cell(row, 3)
    print(cell.value)


print('\n\n\n')

# Getting all rows of cell 3 but excluding the first row
for row in range(2, trasnactions.max_row + 1):
    cell = trasnactions.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = trasnactions.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(trasnactions, min_row=2, max_row=trasnactions.max_row, min_col=4, max_col=4)

chart = BarChart()
chart.add_data(values)

trasnactions.add_chart(chart, "e2")

wb.save('transaction2.xlsx')

""" ## Openpyxl package to work with Excel sheets
import openpyxl as xls
from openpyxl.chart import BarChart, Reference

def process_workbook(filename)
    wb = xls.load_workbook(filename)
    trasnactions = wb['transaction1']

    # Getting all rows of cell 3 
    for row in range(1, trasnactions.max_row + 1):
        cell = trasnactions.cell(row, 3)
        print(cell.value)


    print('\n\n\n')

    # Getting all rows of cell 3 but excluding the first row
    for row in range(2, trasnactions.max_row + 1):
        cell = trasnactions.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = trasnactions.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(trasnactions, min_row=2, max_row=trasnactions.max_row, min_col=4, max_col=4)

    chart = BarChart()
    chart.add_data(values)

    trasnactions.add_chart(chart, "e2")

    wb.save(filename)


process_workbook('transcation.xlsx') """